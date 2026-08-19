"""東京都内、複数区の教育機関一覧(区立小中学校)を、フロントエンド用の1つの
GeoJSONにまとめる。

出典: 各区(多くは「自治体標準データセット」形式)、東京都オープンデータカタログ
      サイト経由(https://catalog.data.metro.tokyo.lg.jp/)
ライセンス: 各区とも CC BY 4.0 (https://creativecommons.org/licenses/by/4.0/deed.ja)

対応区(2026-08時点、15区): 千代田区・中央区・新宿区・台東区・墨田区・江東区・
品川区・世田谷区・中野区・練馬区・板橋区・北区・荒川区・葛飾区・大田区

未対応(見送り):
- 杉並区: カタログ掲載のCSV URLが404(区サイト側でファイルが移動/削除された
  模様)。wagmap版(公共施設情報「学校」)も404で、正しいURLは未特定
- 文京区・渋谷区・豊島区・足立区: カタログ上で「学校」を含むデータセットが
  見つからず(未公開か、この検索語では見つからない名称の可能性)
- 江戸川区: 「小中学校通学区域情報」はあるが、学校の緯度経度ではなく
  通学区域(住所範囲)のデータのため今回の用途には使えない
これらは区サイトを直接確認するか、新しいURL・データセット名が分かれば
`WARDS`に追加できる。

自治体標準データセット形式(教育機関一覧)は区によって細部が違う:
- 文字コード: utf-8-sig / cp932 / utf-16-le のいずれか(区ごとにバラバラ)
- クォート: フィールドを""で囲む区と囲まない区がある(csvモジュールは両対応)
- 学校種コード: A1=幼稚園, B1=小学校, C1=中学校 など(通学路の趣旨に合わせ
  B1/C1のみを採用する)
品川区・中野区・北区・大田区は教育機関一覧と別の独自形式(公共施設POI形式や
区独自の一覧表)を公開しているため、それぞれ個別のパーサーで対応する。
大田区のみXLSX形式(要openpyxl: `uv run --with openpyxl python ...`)。

再取得する場合: opendata/raw_cache/<ward>.csv(またはxlsx) を差し替えてから
    uv run --with openpyxl python opendata/prepare_tokyo_schools.py
"""
import csv
import io
import json
import os
import re

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RAW_DIR = os.path.join(BASE_DIR, "raw_cache")
OUTPUT_PATHS = [
    os.path.join(BASE_DIR, "tokyo_schools.geojson"),
    os.path.join(BASE_DIR, "..", "frontend", "schools.geojson"),
]

# ward: (ファイル名, 出典URL)
WARDS = {
    "千代田区": ("chiyoda.csv", "https://catalog.data.metro.tokyo.lg.jp/dataset/t131016d0000000007"),
    "中央区":   ("chuo.csv", "https://catalog.data.metro.tokyo.lg.jp/dataset/t131024d0000000026"),
    "新宿区":   ("shinjuku.csv", "https://catalog.data.metro.tokyo.lg.jp/dataset/t131041d0000000127"),
    "台東区":   ("taito.csv", "https://catalog.data.metro.tokyo.lg.jp/dataset/t131067d0000000386"),
    "墨田区":   ("sumida.csv", "https://catalog.data.metro.tokyo.lg.jp/dataset/t131075d0000000139"),
    "江東区":   ("koto.csv", "https://catalog.data.metro.tokyo.lg.jp/dataset/t131083d3100000006"),
    "品川区":   ("shinagawa.csv", "https://catalog.data.metro.tokyo.lg.jp/dataset/t131091d0000000160"),
    "世田谷区": ("setagaya.csv", "https://catalog.data.metro.tokyo.lg.jp/dataset/t131121d0000000010"),
    "練馬区":   ("nerima.csv", "https://catalog.data.metro.tokyo.lg.jp/dataset/t131202d0000000126"),
    "板橋区":   ("itabashi.csv", "https://catalog.data.metro.tokyo.lg.jp/dataset/t131199d0000000007"),
    "荒川区":   ("arakawa.csv", "https://catalog.data.metro.tokyo.lg.jp/dataset/t131181d0000000007"),
    "葛飾区":   ("katsushika.csv", "https://catalog.data.metro.tokyo.lg.jp/dataset/t131229d0000000011"),
    "中野区":   ("nakano.csv", "https://catalog.data.metro.tokyo.lg.jp/dataset/t131148d0000000054"),
    "北区":     (("kita_elem.csv", "kita_jhs.csv"), "https://catalog.data.metro.tokyo.lg.jp/dataset/t131172d0000000003"),
    "大田区":   (("ota_elem.xlsx", "ota_jhs.xlsx"), "https://catalog.data.metro.tokyo.lg.jp/dataset/t131113d0000000011"),
}

# 台東区のCSVは、番地(例: "8-16")がExcelで日付("8月16日")に自動変換された
# まま公開されている(原データの不備)。パターンが一定なので元表記に戻す。
_EXCEL_DATE_ADDRESS = re.compile(r"(\d+)月(\d+)日")


def fix_excel_date_address(address):
    return _EXCEL_DATE_ADDRESS.sub(r"\1-\2", address)


TYPE_CODE_MAP = {"B1": "小学校", "C1": "中学校"}

# 千代田区の教育機関一覧CSVは、小中学校(B1/C1)の緯度経度が空欄で、
# 幼稚園(A1)にしか座標が入っていない(原データの不備)。10校のみのため、
# 国土地理院ジオコーディングAPI(住所検索)で個別に座標を補完し、結果を
# ここに固定値として埋め込む(再取得手順はREADME参照。API呼び出しを
# 毎回の実行に含めると、原データ側の一時的な不調で再現性が壊れるため)。
CHIYODA_GEOCODED = [
    {"name": "麹町小学校", "type": "小学校", "address": "麹町2丁目8番地", "postal_code": "1020083", "lat": 35.685356, "lon": 139.739761},
    {"name": "九段小学校", "type": "小学校", "address": "三番町16番地", "postal_code": "1020075", "lat": 35.690456, "lon": 139.740402},
    {"name": "番町小学校", "type": "小学校", "address": "六番町8番地", "postal_code": "1020085", "lat": 35.687614, "lon": 139.732788},
    {"name": "富士見小学校", "type": "小学校", "address": "富士見1丁目10番3号", "postal_code": "1020071", "lat": 35.697475, "lon": 139.746658},
    {"name": "お茶の水小学校", "type": "小学校", "address": "神田猿楽町1丁目1番1号", "postal_code": "1010064", "lat": 35.697388, "lon": 139.760391},
    {"name": "千代田小学校", "type": "小学校", "address": "神田司町2丁目16番地", "postal_code": "1010048", "lat": 35.693047, "lon": 139.768494},
    {"name": "昌平小学校", "type": "小学校", "address": "外神田3丁目4番7号", "postal_code": "1010021", "lat": 35.701454, "lon": 139.769836},
    {"name": "和泉小学校", "type": "小学校", "address": "神田和泉町1番地", "postal_code": "1010024", "lat": 35.699192, "lon": 139.775970},
    {"name": "麹町中学校", "type": "中学校", "address": "平河町2丁目5番1号", "postal_code": "1020093", "lat": 35.680115, "lon": 139.739044},
    {"name": "神田一橋中学校", "type": "中学校", "address": "一ツ橋2丁目6番14号", "postal_code": "1010003", "lat": 35.694248, "lon": 139.756546},
]


def read_text(path):
    raw = open(path, "rb").read()
    for enc in ("utf-8-sig", "cp932", "utf-16-le", "utf-8"):
        try:
            text = raw.decode(enc)
            if "学校" in text or "施設名称" in text:
                return text
        except UnicodeDecodeError:
            continue
    raise ValueError(f"デコードできませんでした: {path}")


def parse_standard_schema(text, ward):
    reader = csv.DictReader(io.StringIO(text))
    out = []
    for r in reader:
        type_code = (r.get("教育機関_学校種") or "").strip()
        school_type = TYPE_CODE_MAP.get(type_code)
        if school_type is None:
            continue
        name = (r.get("教育機関_学校名") or "").strip()
        lat_s = (r.get("教育機関_緯度") or "").strip()
        lon_s = (r.get("教育機関_経度") or "").strip()
        if not name or not lat_s or not lon_s:
            continue
        out.append({
            "name": name,
            "type": school_type,
            "ward": (r.get("教育機関_学校所在地（市区町村）") or ward).strip(),
            "address": (r.get("教育機関_学校所在地（町字）") or "") + (r.get("教育機関_学校所在地（番地以下）") or ""),
            "postal_code": (r.get("教育機関_郵便番号") or "").strip(),
            "lat": float(lat_s),
            "lon": float(lon_s),
        })
    return out


def parse_shinagawa(text):
    """品川区だけ列構成が違う: 施設名称,区,町名以降,郵便番号,電話番号,経度,緯度,分類"""
    reader = csv.DictReader(io.StringIO(text))
    out = []
    for r in reader:
        category = (r.get("分類") or "").strip()
        if category not in ("小学校", "中学校"):
            continue
        name = (r.get("施設名称") or "").strip()
        lat_s = (r.get("緯度") or "").strip()
        lon_s = (r.get("経度") or "").strip()
        if not name or not lat_s or not lon_s:
            continue
        out.append({
            "name": name,
            "type": category,
            "ward": (r.get("区") or "品川区").strip(),
            "address": (r.get("町名以降") or "").strip(),
            "postal_code": (r.get("郵便番号") or "").strip(),
            "lat": float(lat_s),
            "lon": float(lon_s),
        })
    return out


def parse_nakano(text):
    """中野区は公共施設POI形式: 経度,緯度,分類(すべて"小中学校"に統一されて
    おり区別できないため、名称の末尾で小学校/中学校を判定する)。"""
    reader = csv.DictReader(io.StringIO(text))
    out = []
    for r in reader:
        name = (r.get("名称") or "").strip()
        if name.endswith("小学校"):
            school_type = "小学校"
        elif name.endswith("中学校"):
            school_type = "中学校"
        else:
            continue
        lat_s = (r.get("緯度") or "").strip()
        lon_s = (r.get("経度") or "").strip()
        if not lat_s or not lon_s:
            continue
        out.append({
            "name": name,
            "type": school_type,
            "ward": "中野区",
            "address": (r.get("住所") or "").strip(),
            "postal_code": (r.get("郵便番号") or "").strip(),
            "lat": float(lat_s),
            "lon": float(lon_s),
        })
    return out


def parse_kita(elem_text, jhs_text):
    """北区: 施設名,住所,緯度,経度,電話番号,カテゴリ の区独自形式。
    小学校・中学校が別ファイルであること以外は共通。"""
    out = []
    for text, school_type in ((elem_text, "小学校"), (jhs_text, "中学校")):
        reader = csv.DictReader(io.StringIO(text))
        for r in reader:
            name = (r.get("施設名") or "").strip()
            lat_s = (r.get("緯度") or "").strip()
            lon_s = (r.get("経度") or "").strip()
            if not name or not lat_s or not lon_s:
                continue
            out.append({
                "name": name,
                "type": school_type,
                "ward": "北区",
                "address": (r.get("住所") or "").strip(),
                "postal_code": "",
                "lat": float(lat_s),
                "lon": float(lon_s),
            })
    return out


def parse_ota(elem_path, jhs_path):
    """大田区: XLSX形式、2行目がヘッダー(名称,電話,ＦＡＸ,〒,所在地,緯度,経度)。"""
    import openpyxl

    out = []
    for path, school_type in ((elem_path, "小学校"), (jhs_path, "中学校")):
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        header = [c.value for c in ws[2]]
        idx = {name: i for i, name in enumerate(header) if name}
        for row in ws.iter_rows(min_row=3, values_only=True):
            name = row[idx["名称"]]
            if not name:
                continue
            lat = row[idx["緯度"]]
            lon = row[idx["経度"]]
            if lat is None or lon is None:
                continue
            postal = str(row[idx["〒"]]) if row[idx["〒"]] else ""
            out.append({
                "name": str(name).strip(),
                "type": school_type,
                "ward": "大田区",
                "address": str(row[idx["所在地"]] or "").strip(),
                "postal_code": postal,
                "lat": float(lat),
                "lon": float(lon),
            })
    return out


def main():
    all_schools = []
    ward_counts = {}
    for ward, (filename, source_url) in WARDS.items():
        if ward == "千代田区":
            # CSV側は小中学校の緯度経度が空のため、ジオコーディング済みの固定値を使う
            schools = [dict(s, ward="千代田区") for s in CHIYODA_GEOCODED]
        elif ward == "品川区":
            schools = parse_shinagawa(read_text(os.path.join(RAW_DIR, filename)))
        elif ward == "中野区":
            schools = parse_nakano(read_text(os.path.join(RAW_DIR, filename)))
        elif ward == "北区":
            elem_file, jhs_file = filename
            schools = parse_kita(
                read_text(os.path.join(RAW_DIR, elem_file)),
                read_text(os.path.join(RAW_DIR, jhs_file)),
            )
        elif ward == "大田区":
            elem_file, jhs_file = filename
            schools = parse_ota(
                os.path.join(RAW_DIR, elem_file),
                os.path.join(RAW_DIR, jhs_file),
            )
        else:
            schools = parse_standard_schema(read_text(os.path.join(RAW_DIR, filename)), ward)
        for s in schools:
            s["source_url"] = source_url
        all_schools.extend(schools)
        ward_counts[ward] = len(schools)
        print(f"{ward}: {len(schools)}件")

    features = []
    for s in all_schools:
        features.append({
            "type": "Feature",
            "geometry": {"type": "Point", "coordinates": [s["lon"], s["lat"]]},
            "properties": {
                "name": s["name"],
                "type": s["type"],
                "ward": s["ward"],
                "address": fix_excel_date_address(s["address"]),
                "postal_code": s["postal_code"],
            }
        })

    geojson = {
        "type": "FeatureCollection",
        "metadata": {
            "title": "東京都内(15区)の区立小中学校一覧",
            "source": "各区(自治体標準データセット等)、東京都オープンデータカタログサイト経由",
            "license": "CC BY 4.0",
            "license_url": "https://creativecommons.org/licenses/by/4.0/deed.ja",
            "ward_counts": ward_counts,
            "total": len(features),
        },
        "features": features,
    }

    for path in OUTPUT_PATHS:
        with open(path, "w", encoding="utf-8") as f:
            json.dump(geojson, f, ensure_ascii=False, indent=1)
        print(f"合計{len(features)}件 -> {path}")


if __name__ == "__main__":
    main()
